"""
Gera os 5 templates Excel para carga em SharePoint Lists do SBN.
Cada arquivo tem:
- Aba "Data": colunas + dados fake
- Aba "_Schema": dicionario de dados (coluna, tipo SP, required, descricao, exemplo)
- Aba "_README" no primeiro arquivo: instrucoes gerais

Estilo visual: header dourado #c4a265 sobre preto, body legivel.
"""
import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"C:\Users\rodri\OneDrive - Nerpa\JS\SBN\SharePoint_Templates"
os.makedirs(OUT_DIR, exist_ok=True)

# ============== STYLES ==============
GOLD = "C4A265"
DARK = "1A1A1A"
SOFT_BG = "F5F0EB"
LIGHT_BORDER = "D4B87A"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=DARK)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
SCHEMA_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
SCHEMA_HEADER_FILL = PatternFill("solid", fgColor=GOLD)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="000000")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 26


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
        ws.row_dimensions[r].height = 18


def autosize(ws, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            longest = max(longest, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[letter].width = longest


def add_schema_sheet(wb, fields):
    """fields = list of dicts: name, type, required, description, example"""
    ws = wb.create_sheet("_Schema")
    headers = ["Column", "SharePoint Type", "Required", "Description", "Example / Allowed Values"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = SCHEMA_HEADER_FONT
        cell.fill = SCHEMA_HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 26
    for i, f in enumerate(fields, 2):
        ws.cell(row=i, column=1, value=f["name"])
        ws.cell(row=i, column=2, value=f["type"])
        ws.cell(row=i, column=3, value="Yes" if f.get("required") else "No")
        ws.cell(row=i, column=4, value=f.get("description", ""))
        ws.cell(row=i, column=5, value=f.get("example", ""))
        for c in range(1, 6):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = LEFT
        ws.row_dimensions[i].height = 32
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 50
    ws.freeze_panes = "A2"
    return ws


def add_readme_sheet(wb, title, instructions):
    ws = wb.create_sheet("_README", 0)
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=DARK)
    ws.row_dimensions[1].height = 30
    ws["A2"] = "Seminar by Nomination - L'Oreal LATAM | Power Apps Migration"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="666666")
    for i, line in enumerate(instructions, 4):
        ws.cell(row=i, column=1, value=line).font = Font(name="Calibri", size=11)
        ws.row_dimensions[i].height = 20
    ws.column_dimensions["A"].width = 110
    return ws


# ============== DATA: REAL FROM MVP ==============

ZONES_COUNTRIES = ["Brazil", "Mexico", "Argentina", "Colombia", "Chile"]
COUNTRY_DIST = {"Brazil": 13, "Mexico": 11, "Argentina": 8, "Colombia": 7, "Chile": 6}

FIRST_NAMES = ["Ana", "Carlos", "Mariana", "Pedro", "Juliana", "Rafael", "Beatriz", "Lucas",
               "Camila", "Gabriel", "Isabella", "Thiago", "Fernanda", "Diego", "Larissa",
               "Matheus", "Amanda", "Bruno", "Leticia", "Gustavo", "Valentina", "Andre",
               "Sophia", "Ricardo", "Helena", "Felipe", "Maria", "Eduardo", "Laura",
               "Henrique", "Paula", "Joao", "Luiza", "Daniel", "Clara", "Rodrigo", "Alice",
               "Marcelo", "Giovanna", "Vinicius", "Renata", "Tomas", "Sofia", "Javier", "Patricia"]

LAST_NAMES = ["Silva", "Santos", "Oliveira", "Souza", "Costa", "Pereira", "Rodrigues",
              "Almeida", "Nascimento", "Lima", "Araujo", "Fernandes", "Carvalho", "Gomes",
              "Martins", "Rocha", "Ribeiro", "Moreira", "Barros", "Freitas",
              "Garcia", "Lopez", "Hernandez", "Gonzalez", "Perez", "Sanchez"]

TITLES = ["Brand Manager", "Marketing Director", "Regional Manager", "Operations Lead",
          "Finance Manager", "HR Business Partner", "Digital Strategist",
          "Supply Chain Manager", "Product Manager", "Commercial Director",
          "Innovation Lead", "Trade Marketing Manager", "Key Account Manager",
          "Category Manager", "Retail Excellence Manager", "Country Operations Director",
          "Country Division General Manager", "Marketing Mger/ Dir. (Brand/Country)",
          "Brand Commercial Manager/Director"]

MANAGER_NAMES = ["Roberto Mendes", "Patricia Aguilar", "Sebastian Castro", "Maria Elena Vargas",
                 "Diego Fernandes", "Ana Lucia Tavares", "Carlos Eduardo Pinto",
                 "Mariana Reyes", "Felipe Andrade", "Cecilia Romero", "Luis Henrique Costa",
                 "Beatriz Salazar"]

DIVISIONS = ["Luxe", "Consumer Products", "Professional Products", "Active Cosmetics",
             "Operations", "Finance", "HR", "Digital"]
ORG_DIVISIONS = ["GPP", "PLP", "GPPCORP", "CAP", "PPP", "CPP"]
TALENT_CLASSES = ["Future Leaders", "Rising Players", "Essential Players"]
PERFORMANCE = ["Exceeds Expectations", "Strong Performer", "Meets Expectations", "Developing"]
RISK = ["High", "Medium", "Low"]

COURSES = [
    {"id": 1, "name": "Strategic Leadership Accelerator", "code": "SBN-SLA", "investment": 12500,
     "seats": 25, "category": "Leadership", "format": "In-Person", "duration": "5 days",
     "provider": "INSEAD",
     "description": "Advanced strategic leadership program for future executives.",
     "objective": "Develop strategic vision and transformational leadership capabilities.",
     "skills": "Strategic Thinking, Change Management, Executive Presence",
     "minMonthsCompany": 24, "minMonthsDivision": 12, "maxAge": 45,
     "requiredCourse": "", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 2, "name": "Digital Transformation Mastery", "code": "SBN-DTM", "investment": 9800,
     "seats": 30, "category": "Digital", "format": "In-Person", "duration": "4 days",
     "provider": "MIT Sloan",
     "description": "Immersion in digital transformation and technological innovation.",
     "objective": "Empower leaders to drive digital transformation initiatives.",
     "skills": "Digital Strategy, AI & Analytics, Innovation",
     "minMonthsCompany": 18, "minMonthsDivision": 6, "maxAge": 50,
     "requiredCourse": "", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 3, "name": "Global Marketing Excellence", "code": "SBN-GME", "investment": 11200,
     "seats": 20, "category": "Marketing", "format": "In-Person", "duration": "4 days",
     "provider": "HEC Paris",
     "description": "Excellence program in global marketing and brand management.",
     "objective": "Deepen strategic marketing competencies on a global scale.",
     "skills": "Brand Strategy, Consumer Insights, Market Analysis",
     "minMonthsCompany": 36, "minMonthsDivision": 18, "maxAge": 42,
     "requiredCourse": "SBN-DTM", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 4, "name": "Financial Acumen for Leaders", "code": "SBN-FAL", "investment": 8500,
     "seats": 35, "category": "Finance", "format": "In-Person", "duration": "3 days",
     "provider": "London Business School",
     "description": "Financial acumen development for decision-making.",
     "objective": "Strengthen financial analysis and P&L management capabilities.",
     "skills": "Financial Analysis, P&L Management, Business Planning",
     "minMonthsCompany": 12, "minMonthsDivision": 6, "maxAge": 55,
     "requiredCourse": "", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 5, "name": "People Leadership Program", "code": "SBN-PLP", "investment": 7800,
     "seats": 40, "category": "People", "format": "In-Person", "duration": "3 days",
     "provider": "CCL",
     "description": "Program focused on people leadership and team development.",
     "objective": "Develop coaching, feedback, and talent management skills.",
     "skills": "Coaching, Team Building, Talent Development",
     "minMonthsCompany": 12, "minMonthsDivision": 6, "maxAge": 50,
     "requiredCourse": "", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 6, "name": "Supply Chain Innovation", "code": "SBN-SCI", "investment": 9200,
     "seats": 25, "category": "Operations", "format": "In-Person", "duration": "4 days",
     "provider": "IMD",
     "description": "Innovation in supply chain and global operations.",
     "objective": "Transform the supply chain with technology and agility.",
     "skills": "Supply Chain Optimization, Sustainability, Operations Strategy",
     "minMonthsCompany": 24, "minMonthsDivision": 12, "maxAge": 48,
     "requiredCourse": "", "divMin": 12, "divMax": 60, "status": "Active"},
    {"id": 7, "name": "Sustainability & ESG Leadership", "code": "SBN-ESG", "investment": 8900,
     "seats": 30, "category": "ESG", "format": "In-Person", "duration": "3 days",
     "provider": "Cambridge Judge",
     "description": "Leadership in sustainability and ESG practices.",
     "objective": "Integrate ESG into business strategy and operations.",
     "skills": "ESG Strategy, Sustainable Innovation, Stakeholder Management",
     "minMonthsCompany": 18, "minMonthsDivision": 6, "maxAge": 55,
     "requiredCourse": "", "divMin": "", "divMax": "", "status": "Active"},
    {"id": 8, "name": "Luxury Brand Management", "code": "SBN-LBM", "investment": 14000,
     "seats": 15, "category": "Marketing", "format": "In-Person", "duration": "5 days",
     "provider": "Bocconi",
     "description": "Luxury brand management and premium experience.",
     "objective": "Master the codes of luxury and create memorable brand experiences.",
     "skills": "Luxury Marketing, Brand Equity, Premium Positioning",
     "minMonthsCompany": 36, "minMonthsDivision": 24, "maxAge": 40,
     "requiredCourse": "SBN-GME", "divMin": "", "divMax": "", "status": "Active"},
]

COUNTRY_ALLOCS = [
    {"country": "Brazil", "seats": 30, "budget": 255000, "zone": "LATAM"},
    {"country": "Mexico", "seats": 25, "budget": 212500, "zone": "LATAM"},
    {"country": "Argentina", "seats": 20, "budget": 170000, "zone": "LATAM"},
    {"country": "Colombia", "seats": 15, "budget": 127500, "zone": "LATAM"},
    {"country": "Chile", "seats": 10, "budget": 85000, "zone": "LATAM"},
]


# ============== EMPLOYEE GENERATION ==============
random.seed(42)


def generate_employees():
    employees = []
    pre_course_pool = ["", "", "", "", "SBN-DTM", "SBN-PLP", "SBN-FAL", "SBN-SLA", "SBN-GME"]
    eid = 1
    for country in ZONES_COUNTRIES:
        n = COUNTRY_DIST[country]
        for _ in range(n):
            fn = random.choice(FIRST_NAMES)
            ln = random.choice(LAST_NAMES)
            age = random.randint(28, 52)
            mc = random.randint(6, 144)
            md = min(mc, random.randint(3, 60))
            mr = min(md, random.randint(2, 36))
            past_sbn = random.choices([0, 1, 2, 3], weights=[55, 28, 12, 5])[0]
            completed = random.choice(pre_course_pool)
            carol_id = random.randint(10000, 9999999)
            employees.append({
                "CarolID": carol_id,
                "FullName": f"{fn} {ln}",
                "FirstName": fn,
                "LastName": ln,
                "Gender": random.choice(["F", "F", "M", "M", "F"]),
                "Age": age,
                "JobTitle": random.choice(TITLES),
                "Country": country,
                "Zone": "LATAM",
                "Division": random.choice(DIVISIONS),
                "MonthsCompany": mc,
                "MonthsDivision": md,
                "MonthsRole": mr,
                "TalentClass": random.choices(TALENT_CLASSES, weights=[20, 45, 35])[0],
                "Performance": random.choices(PERFORMANCE, weights=[18, 42, 32, 8])[0],
                "SuccessionPipeline": random.choice([True, False, False, True, False]),
                "PastSBNCount": past_sbn,
                "FlightRisk": random.choices(RISK, weights=[18, 32, 50])[0],
                "CompletedCourses": completed,
                "Email": f"{fn.lower()}.{ln.lower()}@loreal.com",
                "ManagerName": random.choice(MANAGER_NAMES),
                "BPCode": f"BP-{random.randint(1, 12):03d}",
                "KeyPlayer": random.choice([True, False, False, True, False, False]),
                "OrgDivision": random.choice(ORG_DIVISIONS),
            })
            eid += 1
    return employees


EMPLOYEES = generate_employees()


# ============== NOMINATIONS ==============

def calc_score(emp):
    score = 0
    score += {"Future Leaders": 30, "Rising Players": 22, "Essential Players": 15}.get(emp["TalentClass"], 0)
    if emp["SuccessionPipeline"]:
        score += 20
    score += {0: 15, 1: 5}.get(emp["PastSBNCount"], 0)
    score += {"High": 15, "Medium": 8, "Low": 0}.get(emp["FlightRisk"], 0)
    score += {"Exceeds Expectations": 20, "Strong Performer": 14,
              "Meets Expectations": 8, "Developing": 3}.get(emp["Performance"], 0)
    return score


def check_eligibility(emp, course):
    if course["minMonthsCompany"] and emp["MonthsCompany"] < course["minMonthsCompany"]:
        return False
    if course["minMonthsDivision"] and emp["MonthsDivision"] < course["minMonthsDivision"]:
        return False
    if course["maxAge"] and emp["Age"] > course["maxAge"]:
        return False
    if course["requiredCourse"] and course["requiredCourse"] not in (emp["CompletedCourses"] or ""):
        return False
    if course["divMin"] and course["divMax"]:
        if emp["MonthsDivision"] < course["divMin"] or emp["MonthsDivision"] > course["divMax"]:
            return False
    return True


NOMINATIONS = []
NOMINATION_HISTORY = []

NOM_DEFS = [
    (0, 1, "final", 1, "Head", False),
    (1, 2, "zone_validation", 2, "BP", False),
    (2, 3, "country_hrd_validation", None, "BP", False),
    (5, 1, "nominated", 3, "Head", False),
    (6, 4, "final", 1, "BP", False),
    (8, 5, "rejected", None, "BP", False),
    (10, 2, "nominated", None, "Head", False),
    (12, 6, "country_hrd_validation", 4, "BP", False),
    (15, 7, "zone_validation", 2, "BP", False),
    (18, 8, "final", 1, "Head", False),
    (4, 3, "nominated", None, "BP", True),
    (11, 8, "zone_validation", 1, "BP", True),
]

STAGES = ["nominated", "country_hrd_validation", "zone_validation", "final"]
BASE_DATE = datetime(2026, 1, 10)

hist_id = 1
for i, (emp_idx, cid, status, priority, role, force_out) in enumerate(NOM_DEFS):
    emp = EMPLOYEES[emp_idx]
    course = next(c for c in COURSES if c["id"] == cid)
    eligible = check_eligibility(emp, course)
    out_of_target = (not eligible) or force_out
    if force_out and eligible:
        out_of_target = True
    override = "Excecao mapeada pela lideranca - sucessor critico aprovado pela Zone HRD" if out_of_target else ""
    score = calc_score(emp)
    nom_id = f"NOM-{i+1:04d}"
    created = BASE_DATE + timedelta(days=i * 2)
    just = ("Excecao mapeada pela lideranca - perfil critico para o desenvolvimento da zona"
            if out_of_target else "Potencial identificado pela lideranca para desenvolvimento acelerado")
    rejection = "Restricao orcamentaria para o ciclo atual" if status == "rejected" else ""
    NOMINATIONS.append({
        "NominationID": nom_id,
        "CarolID": emp["CarolID"],
        "EmployeeName": emp["FullName"],
        "CourseID": course["id"],
        "CourseCode": course["code"],
        "CourseName": course["name"],
        "Investment": course["investment"],
        "Eligible": eligible,
        "OutOfTarget": out_of_target,
        "OverrideReason": override,
        "Score": score,
        "Status": status,
        "CreatedDate": created.strftime("%Y-%m-%d"),
        "Justification": just,
        "NominatorRole": role,
        "NominatorEmail": "dione@loreal.com",
        "Priority": priority if priority else "",
        "RejectionReason": rejection,
    })
    # History
    NOMINATION_HISTORY.append({
        "HistoryID": f"HST-{hist_id:04d}", "NominationID": nom_id, "Action": "created",
        "UserEmail": "dione@loreal.com", "UserName": "Dione",
        "Timestamp": created.strftime("%Y-%m-%d %H:%M:%S"),
        "Details": (f"Nomeado (Excecao - Fora do Target) para {course['name']}: {override}"
                    if out_of_target else f"Nomeado para {course['name']}"),
    })
    hist_id += 1
    stage_idx = STAGES.index(status) if status != "rejected" else 1
    for s in range(stage_idx):
        ts = created + timedelta(days=(s + 1) * 3)
        NOMINATION_HISTORY.append({
            "HistoryID": f"HST-{hist_id:04d}", "NominationID": nom_id,
            "Action": "status_advanced", "UserEmail": "dione@loreal.com",
            "UserName": "Dione", "Timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
            "Details": f"Status: {STAGES[s]} -> {STAGES[s+1]}",
        })
        hist_id += 1
    if status == "rejected":
        ts = created + timedelta(days=4)
        NOMINATION_HISTORY.append({
            "HistoryID": f"HST-{hist_id:04d}", "NominationID": nom_id,
            "Action": "rejected", "UserEmail": "dione@loreal.com", "UserName": "Dione",
            "Timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
            "Details": f"Rejeitado: {rejection}",
        })
        hist_id += 1
    if priority:
        ts = created + timedelta(days=2)
        NOMINATION_HISTORY.append({
            "HistoryID": f"HST-{hist_id:04d}", "NominationID": nom_id,
            "Action": "priority_changed", "UserEmail": "dione@loreal.com",
            "UserName": "Dione", "Timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
            "Details": f"Prioridade definida: P{priority}",
        })
        hist_id += 1


# ============== WRITERS ==============

def write_workbook(path, header_row, data_rows, schema_fields, readme_lines=None, sheet_name="Data"):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)
    for c, h in enumerate(header_row, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, row=1, cols=len(header_row))
    for r_idx, row in enumerate(data_rows, 2):
        for c_idx, v in enumerate(row, 1):
            if isinstance(v, bool):
                ws.cell(row=r_idx, column=c_idx, value="TRUE" if v else "FALSE")
            else:
                ws.cell(row=r_idx, column=c_idx, value=v)
    style_data_rows(ws, 2, 1 + len(data_rows), len(header_row))
    ws.freeze_panes = "A2"
    autosize(ws)
    add_schema_sheet(wb, schema_fields)
    if readme_lines:
        add_readme_sheet(wb, os.path.basename(path).replace(".xlsx", ""), readme_lines)
    wb.save(path)
    print(f"Wrote {path} ({len(data_rows)} rows)")


# ====== 01 Employees ======
emp_headers = ["CarolID", "FullName", "FirstName", "LastName", "Gender", "Age", "JobTitle",
               "Country", "Zone", "Division", "MonthsCompany", "MonthsDivision", "MonthsRole",
               "TalentClass", "Performance", "SuccessionPipeline", "PastSBNCount", "FlightRisk",
               "CompletedCourses", "Email", "ManagerName", "BPCode", "KeyPlayer", "OrgDivision"]
emp_rows = [[e[h] for h in emp_headers] for e in EMPLOYEES]
emp_schema = [
    {"name": "CarolID", "type": "Number (PK)", "required": True, "description": "Carol ID interno Loreal", "example": "10231, 123045"},
    {"name": "FullName", "type": "Single line text", "required": True, "description": "Nome completo", "example": "Ana Silva"},
    {"name": "FirstName", "type": "Single line text", "required": True, "description": "Primeiro nome (usado no export)", "example": "Ana"},
    {"name": "LastName", "type": "Single line text", "required": True, "description": "Sobrenome (usado no export)", "example": "Silva"},
    {"name": "Gender", "type": "Choice", "required": False, "description": "Genero", "example": "F | M"},
    {"name": "Age", "type": "Number", "required": True, "description": "Idade em anos", "example": "28-52"},
    {"name": "JobTitle", "type": "Single line text", "required": True, "description": "Cargo atual", "example": "Brand Manager"},
    {"name": "Country", "type": "Choice", "required": True, "description": "Pais LATAM", "example": "Brazil | Mexico | Argentina | Colombia | Chile"},
    {"name": "Zone", "type": "Choice", "required": True, "description": "Zona", "example": "LATAM"},
    {"name": "Division", "type": "Choice", "required": True, "description": "Divisao L'Oreal", "example": "Luxe | Consumer Products | Professional Products | Active Cosmetics | Operations | Finance | HR | Digital"},
    {"name": "MonthsCompany", "type": "Number", "required": True, "description": "Tenure total na L'Oreal (meses)", "example": "6-144"},
    {"name": "MonthsDivision", "type": "Number", "required": True, "description": "Tempo na divisao atual (meses)", "example": "3-60"},
    {"name": "MonthsRole", "type": "Number", "required": True, "description": "Tempo no cargo atual (meses)", "example": "2-36"},
    {"name": "TalentClass", "type": "Choice", "required": True, "description": "Classificacao de talento", "example": "Future Leaders | Rising Players | Essential Players"},
    {"name": "Performance", "type": "Choice", "required": True, "description": "Avaliacao de performance", "example": "Exceeds Expectations | Strong Performer | Meets Expectations | Developing"},
    {"name": "SuccessionPipeline", "type": "Yes/No", "required": True, "description": "Esta no pipeline de sucessao?", "example": "TRUE | FALSE"},
    {"name": "PastSBNCount", "type": "Number", "required": True, "description": "Quantas vezes ja participou de SBN", "example": "0, 1, 2, 3"},
    {"name": "FlightRisk", "type": "Choice", "required": True, "description": "Risco de saida da empresa", "example": "High | Medium | Low"},
    {"name": "CompletedCourses", "type": "Single line text", "required": False, "description": "Codigos de cursos SBN ja concluidos (separados por virgula)", "example": "SBN-DTM, SBN-PLP"},
    {"name": "Email", "type": "Single line text", "required": True, "description": "E-mail corporativo", "example": "ana.silva@loreal.com"},
    {"name": "ManagerName", "type": "Single line text", "required": True, "description": "Nome do gestor direto", "example": "Roberto Mendes"},
    {"name": "BPCode", "type": "Single line text", "required": False, "description": "Codigo do Business Partner responsavel", "example": "BP-001"},
    {"name": "KeyPlayer", "type": "Yes/No", "required": True, "description": "Talento estrategico (Key Player)", "example": "TRUE | FALSE"},
    {"name": "OrgDivision", "type": "Choice", "required": True, "description": "Sigla organizacional", "example": "GPP | PLP | GPPCORP | CAP | PPP | CPP"},
]
emp_readme = [
    "Lista MASTER de colaboradores LATAM elegiveis para nomeacao SBN.",
    "",
    "VOLUME DE PRODUCAO: ~3.000 registros. Este template traz 45 amostras (mesma quantidade do MVP) distribuidas entre os 5 paises:",
    "Brasil: 13 | Mexico: 11 | Argentina: 8 | Colombia: 7 | Chile: 6",
    "",
    "INSTRUCOES PARA A LOREAL:",
    "1. Manter exatamente os nomes de coluna (Carol ID, Performance, etc.) - o Power App referencia esses nomes.",
    "2. CarolID e a chave primaria - deve ser unico. Vem do HRIS interno.",
    "3. Campos Choice devem usar EXATAMENTE os valores listados na aba _Schema (case-sensitive).",
    "4. Campos Yes/No: usar TRUE / FALSE em maiusculas.",
    "5. CompletedCourses: lista codigos de cursos SBN previamente concluidos (ex.: 'SBN-DTM, SBN-PLP') - vazio se nenhum.",
    "6. Para popular a Lista no SharePoint, importar via 'Quick Edit' ou Power Automate.",
    "",
    "REGRAS QUE USAM ESTES CAMPOS (validadas pelo MVP, ver CLAUDE.md):",
    "- Eligibility: MonthsCompany, MonthsDivision, Age, CompletedCourses",
    "- Scoring: TalentClass, SuccessionPipeline, PastSBNCount, FlightRisk, Performance",
    "- Filtros/Dashboards: Country, Division, KeyPlayer",
]
write_workbook(os.path.join(OUT_DIR, "01_Employees.xlsx"), emp_headers, emp_rows, emp_schema, emp_readme)


# ====== 02 Courses ======
course_headers = ["CourseID", "Name", "Code", "Investment", "Seats", "Category", "Format",
                  "Duration", "Provider", "Description", "Objective", "Skills",
                  "MinMonthsCompany", "MinMonthsDivision", "MaxAge", "RequiredCourseCode",
                  "DivisionRangeMin", "DivisionRangeMax", "Status"]
course_rows = [[c["id"], c["name"], c["code"], c["investment"], c["seats"], c["category"],
                c["format"], c["duration"], c["provider"], c["description"], c["objective"],
                c["skills"], c["minMonthsCompany"], c["minMonthsDivision"], c["maxAge"],
                c["requiredCourse"], c["divMin"], c["divMax"], c["status"]] for c in COURSES]
course_schema = [
    {"name": "CourseID", "type": "Number (PK)", "required": True, "description": "ID sequencial do curso", "example": "1-8"},
    {"name": "Name", "type": "Single line text", "required": True, "description": "Nome completo", "example": "Strategic Leadership Accelerator"},
    {"name": "Code", "type": "Single line text", "required": True, "description": "Codigo curto unico", "example": "SBN-SLA"},
    {"name": "Investment", "type": "Currency (USD)", "required": True, "description": "Custo por participante", "example": "12500"},
    {"name": "Seats", "type": "Number", "required": True, "description": "Vagas totais no curso", "example": "15-40"},
    {"name": "Category", "type": "Choice", "required": True, "description": "Categoria do programa", "example": "Leadership | Digital | Marketing | Finance | People | Operations | ESG"},
    {"name": "Format", "type": "Choice", "required": True, "description": "Formato de entrega", "example": "In-Person | Hybrid | Online"},
    {"name": "Duration", "type": "Single line text", "required": True, "description": "Duracao descritiva", "example": "5 days"},
    {"name": "Provider", "type": "Single line text", "required": True, "description": "Instituicao parceira", "example": "INSEAD | MIT Sloan | HEC Paris"},
    {"name": "Description", "type": "Multiple lines of text", "required": False, "description": "Descricao do curso", "example": "Advanced strategic leadership program..."},
    {"name": "Objective", "type": "Multiple lines of text", "required": False, "description": "Objetivo de aprendizagem", "example": "Develop strategic vision..."},
    {"name": "Skills", "type": "Single line text", "required": False, "description": "Skills desenvolvidas (separadas por virgula)", "example": "Strategic Thinking, Change Management"},
    {"name": "MinMonthsCompany", "type": "Number", "required": False, "description": "Pre-req: tenure minimo na empresa", "example": "24"},
    {"name": "MinMonthsDivision", "type": "Number", "required": False, "description": "Pre-req: tenure minimo na divisao", "example": "12"},
    {"name": "MaxAge", "type": "Number", "required": False, "description": "Pre-req: idade maxima", "example": "45"},
    {"name": "RequiredCourseCode", "type": "Single line text", "required": False, "description": "Pre-req: curso anterior obrigatorio", "example": "SBN-DTM"},
    {"name": "DivisionRangeMin", "type": "Number", "required": False, "description": "Pre-req: faixa minima de tempo na divisao", "example": "12"},
    {"name": "DivisionRangeMax", "type": "Number", "required": False, "description": "Pre-req: faixa maxima de tempo na divisao", "example": "60"},
    {"name": "Status", "type": "Choice", "required": True, "description": "Curso ativo no ciclo atual?", "example": "Active | Inactive"},
]
course_readme = [
    "Catalogo dos 8 seminarios SBN para o ciclo 2026.",
    "",
    "INSTRUCOES:",
    "- Dados ja populados conforme o MVP validado pela Loreal.",
    "- Loreal pode ajustar nomes, valores e pre-requisitos antes de importar para o SharePoint.",
    "- Vazio em pre-reqs = sem restricao naquela dimensao.",
    "- Campos vazios em DivisionRangeMin/Max significam que nao ha restricao de faixa especifica.",
    "",
    "PRE-REQUISITOS - LOGICA (referencia para o Power App):",
    "Um colaborador e ELEGIVEL se:",
    "  MonthsCompany >= MinMonthsCompany E",
    "  MonthsDivision >= MinMonthsDivision E",
    "  Age <= MaxAge E",
    "  RequiredCourseCode em Employees.CompletedCourses (se definido) E",
    "  MonthsDivision dentro de [DivisionRangeMin, DivisionRangeMax] (se definido).",
    "",
    "Se inelegivel, ainda pode ser nomeado como OUT-OF-TARGET com justificativa (OverrideReason).",
]
write_workbook(os.path.join(OUT_DIR, "02_Courses.xlsx"), course_headers, course_rows, course_schema, course_readme)


# ====== 03 CountryAllocations ======
alloc_headers = ["Country", "Seats", "Budget", "Zone"]
alloc_rows = [[a["country"], a["seats"], a["budget"], a["zone"]] for a in COUNTRY_ALLOCS]
alloc_schema = [
    {"name": "Country", "type": "Choice (PK)", "required": True, "description": "Pais LATAM", "example": "Brazil | Mexico | Argentina | Colombia | Chile"},
    {"name": "Seats", "type": "Number", "required": True, "description": "Quota de vagas SBN para o pais", "example": "10-30"},
    {"name": "Budget", "type": "Currency (USD)", "required": True, "description": "Orcamento total para SBN no pais", "example": "85000-255000"},
    {"name": "Zone", "type": "Choice", "required": True, "description": "Zona", "example": "LATAM"},
]
alloc_readme = [
    "Alocacao de vagas e orcamento por pais para o ciclo SBN 2026.",
    "",
    "TOTAIS LATAM: 100 vagas | USD 850.000",
    "",
    "REGRAS:",
    "- Seats = HARD LIMIT. Power App deve avisar quando proximo do limite.",
    "- Budget = SOFT LIMIT. Mostra utilizacao mas nao bloqueia (a Loreal valida exceptions).",
    "- Distribuicao por curso e fluida (nao ha quota por curso x pais).",
]
write_workbook(os.path.join(OUT_DIR, "03_CountryAllocations.xlsx"), alloc_headers, alloc_rows, alloc_schema, alloc_readme)


# ====== 04 Nominations ======
nom_headers = ["NominationID", "CarolID", "EmployeeName", "CourseID", "CourseCode", "CourseName",
               "Investment", "Eligible", "OutOfTarget", "OverrideReason", "Score", "Status",
               "CreatedDate", "Justification", "NominatorRole", "NominatorEmail", "Priority",
               "RejectionReason"]
nom_rows = [[n[h] for h in nom_headers] for n in NOMINATIONS]
nom_schema = [
    {"name": "NominationID", "type": "Single line text (PK)", "required": True, "description": "ID da nomeacao", "example": "NOM-0001"},
    {"name": "CarolID", "type": "Lookup -> Employees", "required": True, "description": "Colaborador nomeado", "example": "10231"},
    {"name": "EmployeeName", "type": "Single line text", "required": False, "description": "Nome desnormalizado (cache p/ performance)", "example": "Ana Silva"},
    {"name": "CourseID", "type": "Lookup -> Courses", "required": True, "description": "Curso da nomeacao", "example": "1"},
    {"name": "CourseCode", "type": "Single line text", "required": False, "description": "Codigo curso desnormalizado", "example": "SBN-SLA"},
    {"name": "CourseName", "type": "Single line text", "required": False, "description": "Nome curso desnormalizado", "example": "Strategic Leadership Accelerator"},
    {"name": "Investment", "type": "Currency (USD)", "required": True, "description": "Valor copiado do curso", "example": "12500"},
    {"name": "Eligible", "type": "Yes/No", "required": True, "description": "Resultado do checkEligibility", "example": "TRUE | FALSE"},
    {"name": "OutOfTarget", "type": "Yes/No", "required": True, "description": "Nomeado apesar de inelegivel", "example": "TRUE | FALSE"},
    {"name": "OverrideReason", "type": "Multiple lines of text", "required": False, "description": "Obrigatorio se OutOfTarget=TRUE", "example": "Sucessor critico aprovado pela Zone HRD"},
    {"name": "Score", "type": "Number", "required": True, "description": "Score 0-100 do colaborador (calcScore)", "example": "45-95"},
    {"name": "Status", "type": "Choice", "required": True, "description": "Estagio do workflow", "example": "nominated | country_hrd_validation | zone_validation | final | rejected"},
    {"name": "CreatedDate", "type": "Date", "required": True, "description": "Data de criacao", "example": "2026-01-10"},
    {"name": "Justification", "type": "Multiple lines of text", "required": True, "description": "Texto livre do nomeador", "example": "Potencial identificado..."},
    {"name": "NominatorRole", "type": "Choice", "required": True, "description": "Papel do nomeador", "example": "BP | Head"},
    {"name": "NominatorEmail", "type": "Single line text", "required": True, "description": "Email do nomeador", "example": "dione@loreal.com"},
    {"name": "Priority", "type": "Number", "required": False, "description": "Ranking interno P1-P10 (vazio se nao ranqueado)", "example": "1-10"},
    {"name": "RejectionReason", "type": "Multiple lines of text", "required": False, "description": "Obrigatorio se Status=rejected", "example": "Restricao orcamentaria"},
]
nom_readme = [
    "Nomeacoes ativas no ciclo. Cobre todos os status do workflow.",
    "",
    "AMOSTRAS NESTE TEMPLATE: 12 nomeacoes",
    "- 3 final (aprovadas)",
    "- 3 zone_validation",
    "- 3 country_hrd_validation",
    "- 3 nominated",
    "- 1 rejected",
    "- 2 OutOfTarget (excecao mapeada pela lideranca)",
    "",
    "OBSERVACOES:",
    "- Em producao a lista comeca vazia. Estas amostras existem para testar o app.",
    "- EmployeeName, CourseCode e CourseName sao DESNORMALIZADOS (copiados no momento da nomeacao para evitar joins em runtime).",
    "- Score e congelado no momento da nomeacao (snapshot).",
    "- OverrideReason e obrigatorio quando OutOfTarget=TRUE.",
    "- RejectionReason e obrigatorio quando Status=rejected.",
]
write_workbook(os.path.join(OUT_DIR, "04_Nominations.xlsx"), nom_headers, nom_rows, nom_schema, nom_readme)


# ====== 05 NominationHistory ======
hist_headers = ["HistoryID", "NominationID", "Action", "UserEmail", "UserName", "Timestamp", "Details"]
hist_rows = [[h[k] for k in hist_headers] for h in NOMINATION_HISTORY]
hist_schema = [
    {"name": "HistoryID", "type": "Single line text (PK)", "required": True, "description": "ID do evento", "example": "HST-0001"},
    {"name": "NominationID", "type": "Lookup -> Nominations", "required": True, "description": "Nomeacao referenciada", "example": "NOM-0001"},
    {"name": "Action", "type": "Choice", "required": True, "description": "Tipo de evento", "example": "created | status_advanced | priority_changed | rejected"},
    {"name": "UserEmail", "type": "Single line text", "required": True, "description": "Quem disparou a acao", "example": "dione@loreal.com"},
    {"name": "UserName", "type": "Single line text", "required": True, "description": "Nome humano do usuario", "example": "Dione"},
    {"name": "Timestamp", "type": "Date and Time", "required": True, "description": "Momento exato do evento", "example": "2026-01-10 14:32:00"},
    {"name": "Details", "type": "Multiple lines of text", "required": True, "description": "Descricao humana do que aconteceu", "example": "Status: nominated -> country_hrd_validation"},
]
hist_readme = [
    "Audit trail completo - uma linha por evento por nomeacao.",
    "",
    f"AMOSTRAS NESTE TEMPLATE: {len(NOMINATION_HISTORY)} eventos derivados das 12 nomeacoes.",
    "",
    "EVENTOS:",
    "- created: nomeacao criada (1 por nomeacao)",
    "- status_advanced: avanco de estagio (0-3 por nomeacao)",
    "- priority_changed: prioridade definida/alterada (0+ por nomeacao)",
    "- rejected: nomeacao rejeitada (apenas se Status=rejected)",
    "",
    "OBSERVACOES:",
    "- Em producao a lista comeca vazia e cresce com o uso.",
    "- O Power App / Power Automate deve criar uma linha aqui a cada acao relevante.",
    "- Details e texto livre legivel - usado nas telas de auditoria.",
]
write_workbook(os.path.join(OUT_DIR, "05_NominationHistory.xlsx"), hist_headers, hist_rows, hist_schema, hist_readme)

print("\nDone.")
print(f"Total employees: {len(EMPLOYEES)}")
print(f"Total nominations: {len(NOMINATIONS)}")
print(f"Total history events: {len(NOMINATION_HISTORY)}")
